import random
import json
import os

from pico2d import *
import game_framework
import Title_state
import openpyxl  # 엑셀 사용 라이브러리

name = "Class_files"

grass = None

map = None

mouse_x, mouse_y = 1366 / 2, 768 / 2

count = 1
tmp = 20  # 스프라이트 속도 제한 변수


# CARD 속성
class Card:
    global move_card
    def __init__(self, num):
        # 엑셀로 카드이미지,값처리
        self.wb = openpyxl.load_workbook('Card_list.xlsx', data_only=True)
        self.ws = self.wb['Sheet1']
        self.x, self.y = 400 + 150 * num, 125
        self.layer = 2


# CARD 메서드
class Card_Attack(Card):
    # 상속 받아서 image등록
    card_attack = None

    def __init__(self, num):
        super().__init__(num)
        if Card_Attack.card_attack == None:
            self.image = load_image(self.ws['E2'].value)

    def draw(self):
        if move_card==0:
            self.image.draw(self.x, self.y)

    def update(self):
        if move_card==1:
            self.image.darw(mouse_x, mouse_y)


class Card_Shield(Card):
    # 상속 받아서 image등록
    card_shield = None

    def __init__(self, num):
        super().__init__(num)
        if Card_Shield.card_shield == None:
            self.image = load_image(self.ws['E3'].value)

    def draw(self):
        if move_card==0:
            self.image.draw(self.x, self.y)

    def update(self):
        if move_card==1:
            self. image.darw(mouse_x,mouse_y)

class Grass:
    def __init__(self):
        self.image = load_image('grass2.png')
        self.layer = 0

    def draw(self):
        self.image.draw(683, 150)


class Map:
    def __init__(self):
        self.image = load_image("Map1.png")
        self.layer = 0

    def draw(self):
        self.image.draw(683, 384)


class Tengo:
    tengo = None

    def __init__(self):
        global tmp
        self.x, self.y = 266, 350
        self.frame = 0
        if Tengo.tengo == None:
            self.image = load_image('tengo_sleep.png')

        self.layer = 1

    def update(self):
        global tmp
        self.frame += 1
        if self.frame / tmp >= 12:
            self.frame = 0

    def draw(self):
        self.image.clip_draw((self.frame // tmp) * 200, 0, 200, 200, self.x, self.y)


class Slime:
    slime = None

    def __init__(self, num):
        self.x, self.y = 866 + 100 * num, 300
        self.frame = random.randint(0, 6)
        if Slime.slime == None:
            self.image = load_image('slime_sleep.png')

        self.layer = 2

    def update(self):
        global tmp
        self.frame += 1
        if self.frame / tmp >= 7:
            self.frame = 0

    def draw(self):
        global tmp
        self.image.clip_draw((self.frame // tmp) * 200, 0, 200, 200, self.x, self.y)

move_card=0

def enter():
    global tengo, slimes, grass, map, curser, monster, card
    monster = 0
    curser = load_image('curser.png')
    tengo = Tengo()
    # slimes=[Slime() for monster in range(3)]
    slimes = []  # 슬라임들 리스트 생성
    for i in range(0, random.randint(1,5), 1):  # 슬라임들 시작 끝 스텝순으로
        slimes.append(Slime(i))  # 생성하기

    card = []
    for i in range(0, 5, 1):
        rand = random.randint(1, 2)
        if rand == 1:
            card.append(Card_Attack(i))
        elif rand == 2:
            card.append(Card_Shield(i))
    grass = Grass()
    map = Map()


def exit():
    global tengo, slimes, grass, map, card_shield, curser, card
    del (tengo)
    del (grass)
    del (slimes)  # 궁금!
    del (map)
    del (card)
    del (curser)


def pause():
    pass


def resume():
    pass


def handle_events():
    global count, mouse_x, mouse_y,move_card
    events = get_events()
    for event in events:
        if event.type == SDL_QUIT:
            game_framework.quit()
        elif event.type == SDL_KEYDOWN and event.key == SDLK_ESCAPE:
            game_framework.change_state(Title_state)
        elif event.type == SDL_MOUSEMOTION:
            mouse_x, mouse_y = event.x, 768 - 1 - event.y
        elif event.type == SDL_MOUSEBUTTONDOWN:
            if mouse_x > 400 and mouse_x < 550 and mouse_y > 25 and mouse_y < 225:
                pass

def update():
    global c
    hide_cursor()
    tengo.update()
    for slime in slimes:
        slime.update()
    for c in card:
        c.update()


def draw():
    global mouse_x, mouse_y,c
    clear_canvas()

    map.draw()
    grass.draw()
   # for c in card:
    #    c.draw()
    #tengo.draw()
    for slime in slimes:
        slime.draw()
    curser.draw(mouse_x, mouse_y)

    update_canvas()
